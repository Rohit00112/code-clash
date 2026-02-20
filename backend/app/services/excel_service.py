"""Excel export service - generate reports"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from datetime import datetime
from pathlib import Path
from typing import List

from app.config import settings
from app.models.user import User
from app.models.submission import Submission
from app.services.submission_service import submission_service
from app.services.challenge_loader import challenge_loader
import logging

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for generating Excel reports"""
    
    @staticmethod
    def generate_results_report(db: Session) -> str:
        """
        Generate comprehensive results report
        
        Args:
            db: Database session
            
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        ExcelService._create_leaderboard_sheet(wb, db)
        ExcelService._create_detailed_submissions_sheet(wb, db)
        ExcelService._create_statistics_sheet(wb, db)
        
        # Save file
        exports_dir = Path(settings.get_exports_dir()) / "results"
        exports_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = exports_dir / filename
        
        wb.save(filepath)
        logger.info(f"Generated results report: {filepath}")
        
        return str(filepath)
    
    @staticmethod
    def _create_leaderboard_sheet(wb: Workbook, db: Session):
        """Create leaderboard sheet"""
        ws = wb.create_sheet("Leaderboard")

        # Build dynamic question columns
        questions = challenge_loader.get_available_questions()
        q_headers = [f"Q{q['number']} ({q['title']})" for q in questions]
        q_ids = [q["id"] for q in questions]

        # Headers
        headers = ["Rank", "Username", "Total Score"] + q_headers + ["Avg Time (s)", "Questions Solved"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Get leaderboard data
        leaderboard = submission_service.get_leaderboard(db)

        # Add data
        for entry in leaderboard:
            row = [
                entry["rank"],
                entry["username"],
                entry["total_score"],
            ] + [
                entry["question_scores"].get(q_id, 0) for q_id in q_ids
            ] + [
                entry["avg_execution_time"],
                entry["questions_solved"]
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _create_detailed_submissions_sheet(wb: Workbook, db: Session):
        """Create detailed submissions sheet"""
        ws = wb.create_sheet("Detailed Submissions")
        
        # Headers
        headers = ["Submission ID", "Username", "Question", "Language", "Score", 
                   "Status", "Execution Time", "Submitted At"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get all submissions
        submissions = db.query(Submission).join(User).order_by(
            Submission.submitted_at.desc()
        ).all()
        
        # Add data
        for sub in submissions:
            row = [
                sub.id,
                sub.user.username,
                sub.question_id,
                sub.language,
                sub.score,
                sub.status,
                sub.execution_time,
                sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sub.submitted_at else ""
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _create_statistics_sheet(wb: Workbook, db: Session):
        """Create statistics sheet"""
        ws = wb.create_sheet("Statistics")
        
        # Title
        ws.append(["Code Clash Platform - Event Statistics"])
        ws.append([])
        ws["A1"].font = Font(bold=True, size=14)
        
        # General stats
        total_users = db.query(User).filter(User.role == "participant").count()
        total_submissions = db.query(Submission).count()
        completed_submissions = db.query(Submission).filter(Submission.status == "completed").count()
        
        ws.append(["General Statistics"])
        ws.append(["Total Participants", total_users])
        ws.append(["Total Submissions", total_submissions])
        ws.append(["Completed Submissions", completed_submissions])
        ws.append([])
        
        # Question-wise stats
        ws.append(["Question-wise Statistics"])
        ws.append(["Question", "Total Submissions", "Avg Score", "Max Score"])
        
        from sqlalchemy import func
        
        question_stats = db.query(
            Submission.question_id,
            func.count(Submission.id).label('count'),
            func.avg(Submission.score).label('avg_score'),
            func.max(Submission.score).label('max_score')
        ).filter(
            Submission.status == 'completed'
        ).group_by(
            Submission.question_id
        ).all()
        
        for q_id, count, avg_score, max_score in question_stats:
            ws.append([q_id, count, round(float(avg_score or 0), 2), max_score])
        
        # Style
        for row in ws.iter_rows(min_row=3, max_row=3):
            for cell in row:
                cell.font = Font(bold=True)
        
        for row in ws.iter_rows(min_row=8, max_row=8):
            for cell in row:
                cell.font = Font(bold=True)
    
    @staticmethod
    def generate_credentials_export(users: List[dict]) -> str:
        """
        Generate credentials export for bulk imported users
        
        Args:
            users: List of user credentials
            
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "User Credentials"
        
        # Headers
        headers = ["Username", "Password", "Role", "Login URL"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for user in users:
            row = [
                user["username"],
                user["password"],
                user["role"],
                "http://localhost:8000"  # Update with actual URL
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        exports_dir = Path(settings.get_exports_dir()) / "credentials"
        exports_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"credentials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = exports_dir / filename
        
        wb.save(filepath)
        logger.info(f"Generated credentials export: {filepath}")
        
        return str(filepath)


# Singleton instance
excel_service = ExcelService()
